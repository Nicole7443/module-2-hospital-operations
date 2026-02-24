import gurobipy as gp
from gurobipy import GRB
import openpyxl
import os

# Define a base directory for file paths using the username
username = "Nicol"  # Update this username as needed
base_dir = os.path.join("C:\\Users", username, "Documents\\")

# Update file paths to use the base directory
wb = openpyxl.load_workbook(base_dir + "NICOLE - DO NOT AUTO SAVE! Module 2 - Kate VBA Version.xlsm", keep_vba=True)
sheet = wb['Dept Breakdown']  # Access the specific sheet by name

# 2. Extract Inputs from specific cells
# ER Department (adjust cell references as needed)
er_available_practitioners = sheet['B2'].value 
er_low_severity_arrivals_waiting = sheet['B3'].value 
er_high_severity_arrivals_waiting = sheet['B4'].value 
er_additional_staff_called = sheet['B5'].value 
er_patients_sharing_nurses = sheet['B6'].value  
er_low_severity_patients = sheet['B7'].value  
er_high_severity_patients = sheet['B8'].value 

# Surgery Department
surgery_available_practitioners = sheet['E2'].value  
surgery_low_severity_arrivals_waiting = sheet['E3'].value  
surgery_high_severity_arrivals_waiting = sheet['E4'].value 
surgery_additional_staff_called = sheet['E5'].value 
surgery_request_waiting = sheet['E6'].value  
surgery_low_severity_patients = sheet['E7'].value  
surgery_high_severity_patients = sheet['E8'].value 

# Step-up Department
stepup_available_practitioners = sheet['B11'].value 
stepup_low_severity_arrivals_waiting = sheet['B12'].value 
stepup_high_severity_arrivals_waiting = sheet['B13'].value 
stepup_additional_staff_called = sheet['B14'].value 
stepup_patients_sharing_nurses = sheet['B15'].value 
stepup_request_waiting = sheet['B16'].value 
stepup_low_severity_patients = sheet['B17'].value 
stepup_high_severity_patients = sheet['B18'].value 

# Critical Care Department
critical_available_practitioners = sheet['E11'].value 
critical_low_severity_arrivals_waiting = sheet['E12'].value 
critical_high_severity_arrivals_waiting = sheet['E13'].value 
critical_additional_staff_called = sheet['E14'].value 
critical_request_waiting = sheet['E15'].value 
critical_low_severity_patients = sheet['E16'].value 
critical_high_severity_patients = sheet['E17'].value 

# Initialize Model
model = gp.Model("hospital_opti_model")

# Constants
#depts = [1, 2, 3, 4] # ER, Surgery, Critical Care, Step-down
depts = [1, 2, 3, 4] 
rooms = {1: 25, 2: 9, 3: 18, 4: 30} #rooms in hospital game

# Initial waiting patients by department
initial_low_waiting = {
    1: er_low_severity_arrivals_waiting,
    2: surgery_low_severity_arrivals_waiting,
    3: critical_low_severity_arrivals_waiting,
    4: stepup_low_severity_arrivals_waiting
}

initial_high_waiting = {
    1: er_high_severity_arrivals_waiting,
    2: surgery_high_severity_arrivals_waiting,
    3: critical_high_severity_arrivals_waiting,
    4: stepup_high_severity_arrivals_waiting
}

# Decision Variables
l = model.addVars(depts, vtype=GRB.INTEGER, name="low_priv")
h = model.addVars(depts, vtype=GRB.INTEGER, name="high_priv")
a = model.addVars(depts, vtype=GRB.INTEGER, name="additional_staff")
rl_i = model.addVars(depts, vtype=GRB.INTEGER, name="low_severity_transferred")
rh_i = model.addVars(depts, vtype=GRB.INTEGER, name="high_severity_transferred")

# Nurse sharing only exists for ER (dept 1) and Step-down (dept 4)
sharing_depts = [1, 4]
s_pairs = model.addVars(sharing_depts, vtype=GRB.INTEGER, name="shared_pairs")
s = model.addVars(sharing_depts, vtype=GRB.INTEGER, name="low_shared")

# Available staff per department
available_staff = {
    1: er_available_practitioners,
    2: surgery_available_practitioners,
    3: critical_available_practitioners,
    4: stepup_available_practitioners
}

staff_total = 61 + gp.quicksum(a[i] for i in depts)

#Cost Penalties
pen_arrivals_waiting_ER_c = 150
pen_arrivals_waiting_surgery_c = 3750
pen_arrivals_waiting_critical_care_c = 3750
pen_arrivals_waiting_stepdown_c = 3750

pen_additional_staff_c = 40

#Quality Penalties
pen_waiting_low = 10
pen_waiting_high = 20
pen_sharing_rooms = 5

# Map department-specific cost penalties for waiting
# depts: 1=ER, 2=Surgery, 3=Critical Care, 4=Step-down
pen_waiting_cost = {
    1: pen_arrivals_waiting_ER_c,
    2: pen_arrivals_waiting_surgery_c,
    3: pen_arrivals_waiting_critical_care_c,
    4: pen_arrivals_waiting_stepdown_c
}

# Calculate total cost penalties
# Penalty applied to patients REMAINING in waiting room (not transferred)
cost_penalty = (
    pen_additional_staff_c * gp.quicksum(a[i] for i in depts) +
    gp.quicksum(pen_waiting_cost[i] * ((initial_low_waiting[i] - rl_i[i]) + (initial_high_waiting[i] - rh_i[i])) for i in depts)
)

# Calculate total quality penalties
# Penalty applied to patients REMAINING in waiting room (not transferred)
quality_penalty = (
    pen_sharing_rooms * gp.quicksum(s[i] for i in sharing_depts) +
    gp.quicksum(pen_waiting_low * (initial_low_waiting[i] - rl_i[i]) + pen_waiting_high * (initial_high_waiting[i] - rh_i[i]) for i in depts)
)

# Set objective function to: Min 0.3*(Cost Penalties) + 0.7*(Quality Penalties)
model.setObjective(0.25 * cost_penalty + 0.75 * quality_penalty, GRB.MINIMIZE)

# Constraints

# Ensures total staff allocated does not exceed available staff
# Only ER and Step-down have nurse sharing
model.addConstr(
    gp.quicksum(h[i] + l[i] + rh_i[i] for i in depts) + 
    gp.quicksum(0.5*s[i] for i in sharing_depts) 
    <= staff_total
)

# Ensure s[i] is even (s = 2 * number of pairs)
for i in sharing_depts:
    model.addConstr(s[i] == 2 * s_pairs[i])

for i in depts:
    # Room Capacity - includes current patients and transferred patients
    # s[i] is doubled up, so 0.5 footprint per patient (only for ER and Step-down)
    if i in sharing_depts:
        model.addConstr(h[i] + l[i] + 0.5*s[i] + rh_i[i] <= rooms[i])
    else:
        model.addConstr(h[i] + l[i] + rh_i[i] <= rooms[i])
    
    # Staffing per Department - staff needed must not exceed available staff + additional called in
    if i in sharing_depts:
        model.addConstr(h[i] + l[i] + 0.5*s[i] + rh_i[i] <= available_staff[i] + a[i])
    else:
        model.addConstr(h[i] + l[i] + rh_i[i] <= available_staff[i] + a[i])
    
    # Cannot transfer more patients than are waiting
    model.addConstr(rl_i[i] <= max(0, initial_low_waiting[i]), name=f"max_transfer_low_{i}")
    model.addConstr(rh_i[i] <= max(0, initial_high_waiting[i]), name=f"max_transfer_high_{i}")

model.optimize()

# Access optimal values with .X attribute
new_low_waiting = {}
new_high_waiting = {}
new_low_patients = {}
new_high_patients = {}

# Ensure all outputs are integers with no decimal spots and non-negative
if model.status == GRB.OPTIMAL:
    print(f"Optimal objective value: {int(model.objVal)}")
    for i in depts:
        print(f"Dept {i}: a={int(a[i].X)}, rl_i={int(rl_i[i].X)}, rh_i={int(rh_i[i].X)}")
        
        # Updated waiting room counts
        new_low_waiting[i] = max(0, int(initial_low_waiting[i]) - int(rl_i[i].X))
        new_high_waiting[i] = max(0, int(initial_high_waiting[i]) - int(rh_i[i].X))

        # Updated patients in departments
        new_low_patients[i] = max(0, int(l[i].X + (s[i].X if i in sharing_depts else 0) + rl_i[i].X))
        new_high_patients[i] = max(0, int(h[i].X + rh_i[i].X))

# Add non-negativity constraints for all variables
for i in depts:
    model.addConstr(a[i] >= 0, name=f"non_neg_a_{i}")
    model.addConstr(rl_i[i] >= 0, name=f"non_neg_rl_i_{i}")
    model.addConstr(rh_i[i] >= 0, name=f"non_neg_rh_i_{i}")
    model.addConstr(l[i] >= 0, name=f"non_neg_l_{i}")
    model.addConstr(h[i] >= 0, name=f"non_neg_h_{i}")
    if i in sharing_depts:
        model.addConstr(s[i] >= 0, name=f"non_neg_s_{i}")
        model.addConstr(s_pairs[i] >= 0, name=f"non_neg_s_pairs_{i}")

# Write results back to temp Excel file
file_path = base_dir + "TempOptiData.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb["Opti Results"]

# Clear previous data in the sheet
ws.delete_rows(2, ws.max_row - 1)  # Deletes all rows starting from row 2
ws.delete_cols(4, ws.max_column - 3)  # Deletes all columns starting from column 4

# Add headers for the scores
ws['J1'] = "Total Score"
ws['K1'] = "Cost Score"
ws['L1'] = "Quality Score"

# Write z score, cost score, and quality score to specific cells
z_score = model.objVal  # Assuming z_score is the objective value
cost_score = cost_penalty.getValue()  # Assuming cost_penalty is a Gurobi expression
quality_score = quality_penalty.getValue()  # Assuming quality_penalty is a Gurobi expression

ws['J2'] = z_score
ws['K2'] = cost_score
ws['L2'] = quality_score

# Write additional details to the "Opti Results" sheet
headers = [
    "Department", "Available Practitioners", "Low Severity Arrivals Waiting", "High Severity Arrivals Waiting", 
    "Additional Staff Called in", "Patients Sharing Nurses", "Request Waiting", 
    "Low Severity Patients", "High Severity Patients"
]

# Write headers to the first row
for col_num, header in enumerate(headers, start=1):
    ws.cell(row=1, column=col_num, value=header)

# Write data for each department
row = 2
for i in depts:
    if i == 1:
        dept_name = "ER"
    elif i == 2:
        dept_name = "Surgery"
    elif i == 3:
        dept_name = "Critical Care"
    elif i == 4:
        dept_name = "Step-Up"

    ws.cell(row=row, column=1, value=dept_name)
    ws.cell(row=row, column=2, value=available_staff[i])
    ws.cell(row=row, column=3, value=initial_low_waiting[i])
    ws.cell(row=row, column=4, value=initial_high_waiting[i])
    ws.cell(row=row, column=5, value=int(a[i].X))
    ws.cell(row=row, column=6, value=int(s[i].X) if i in sharing_depts else None)  # Use None instead of "N/A"

    # Handle request waiting values
    if i == 4:
        request_waiting = stepup_request_waiting
    elif i == 3:
        request_waiting = critical_request_waiting
    elif i == 2:
        request_waiting = surgery_request_waiting
    else:
        request_waiting = None

    ws.cell(row=row, column=7, value=request_waiting)
    ws.cell(row=row, column=8, value=new_low_patients[i])
    ws.cell(row=row, column=9, value=new_high_patients[i])
    row += 1

# Write insights to the "Insights" sheet
insights_ws = wb["Insights"]

# Clear previous insights starting from row 2
for row in insights_ws.iter_rows(min_row=2, max_col=1):
    for cell in row:
        cell.value = None

# Map department IDs to names
department_names = {
    1: "ER Department",
    2: "Surgery Department",
    3: "Critical Care Department",
    4: "Step Down Department"
}

# Write insights for each department starting from A2
row = 2
for i in depts:
    dept_name = department_names[i]
    insights_ws[f"A{row}"] = f"{dept_name} had {initial_low_waiting[i]} low severity arrivals and {initial_high_waiting[i]} high severity arrivals."
    row += 1
    insights_ws[f"A{row}"] = f"{dept_name} called {int(a[i].X)} additional nurses."
    row += 1
    insights_ws[f"A{row}"] = f"{dept_name} transferred {int(rl_i[i].X)} low severity patients and {int(rh_i[i].X)} high severity patients."
    row += 1
    insights_ws[f"A{row}"] = f"{dept_name} now has {int(new_low_patients[i])} low severity patients and {int(new_high_patients[i])} high severity patients."
    row += 1

# Save the workbook
wb.save(file_path)
print(f"Successfully wrote scores to 'Opti Results' sheet in {file_path}")